"""
Batimento de Ponto com carga horária diária ajustável (GUI ou modo console)

Alterações:
- Cada usuário tem carga horária padrão (8h) salva em users.json.
- Ao registrar ponto, o usuário pode ajustar a carga horária apenas para aquele dia.
- Planilha salva em .xlsx com horas previstas reais usadas no dia.
- Funciona em GUI (Tkinter) ou modo console se Tkinter não estiver disponível.
"""

import os
import json
from datetime import datetime, date, timedelta, time as dt_time
import locale

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    print("Erro: é necessário instalar openpyxl (pip install openpyxl)")
    raise

try:
    locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
except Exception:
    pass

USERS_FILE = 'users.json'
TEMPO_MINIMO_ENTRE_BATIDAS_MIN = 5
DIAS_PT = ['Segunda-feira','Terça-feira','Quarta-feira','Quinta-feira','Sexta-feira','Sábado','Domingo']
MESES_PT = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho','Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']

def weekday_name_pt(d): return DIAS_PT[d.weekday()]
def month_name_pt(m): return MESES_PT[m-1]

def load_users():
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE,'r',encoding='utf-8') as f: return json.load(f)
    return {}

def save_users(users):
    with open(USERS_FILE,'w',encoding='utf-8') as f: json.dump(users,f,indent=2,ensure_ascii=False)

def get_planilha_path(folder,user):
    return os.path.join(folder,f"batidas_{user.replace(' ','_')}.xlsx")

def ensure_workbook(path):
    if os.path.exists(path): wb=load_workbook(path); ws=wb.active
    else:
        wb=Workbook(); ws=wb.active; ws.title='Batidas'
        ws.append(["Data","Dia da Semana","Mês","Hora 1","Hora 2","Total (h)","Horas Previstas","Saldo (h)"])
    return wb,ws

def parse_time_str(t):
    try:
        h,m=t.split(':'); return dt_time(int(h),int(m))
    except: return None

def safe_combine(now,t): return datetime(now.year,now.month,now.day,t.hour,t.minute)

class PontoCore:
    def __init__(self,planilha_dir=None):
        self.users=load_users()
        self.planilha_dir=planilha_dir or os.path.expanduser('~')

    def register_user(self,username,password):
        if not username: raise ValueError('Usuário vazio')
        self.users[username]={'senha':password,'horas_dia':7.2}
        save_users(self.users)

    def set_user_hours(self,username,horas):
        if username in self.users:
            self.users[username]['horas_dia']=float(horas)
            save_users(self.users)

    def register_punch(self,username,password,horas_dia=None):
        if username not in self.users or self.users[username]['senha']!=password:
            raise PermissionError('Usuário ou senha incorretos')
        horas_previstas=float(horas_dia) if horas_dia else self.users[username].get('horas_dia',7.2)
        now=datetime.now()
        data_str=now.strftime('%d/%m/%Y')
        path=get_planilha_path(self.planilha_dir,username)
        wb,ws=ensure_workbook(path)
        dia=weekday_name_pt(now.date()); mes=month_name_pt(now.month)

        for i,row in enumerate(ws.iter_rows(min_row=2,values_only=True),start=2):
            if row[0]==data_str:
                hora1,hora2=row[3],row[4]
                last=hora2 or hora1
                if last:
                    diff=(now-safe_combine(now,parse_time_str(last))).total_seconds()/60
                    if diff<TEMPO_MINIMO_ENTRE_BATIDAS_MIN:
                        raise RuntimeError('Batida muito próxima da anterior.')
                if not hora1:
                    ws.cell(row=i,column=4,value=now.strftime('%H:%M'))
                elif not hora2:
                    h1=parse_time_str(hora1); total=(now-safe_combine(now,h1)).total_seconds()/3600
                    saldo=total-horas_previstas
                    ws.cell(row=i,column=5,value=now.strftime('%H:%M'))
                    ws.cell(row=i,column=6,value=round(total,2))
                    ws.cell(row=i,column=7,value=horas_previstas)
                    ws.cell(row=i,column=8,value=round(saldo,2))
                else: raise RuntimeError('Duas batidas já feitas.')
                wb.save(path); return path
        ws.append([data_str,dia,mes,now.strftime('%H:%M'),'','',horas_previstas,''])
        wb.save(path); return path

    def close_sheet(self,username):
        path=get_planilha_path(self.planilha_dir,username)
        if not os.path.exists(path): raise FileNotFoundError
        wb=load_workbook(path); ws=wb.active
        total=saldo=0.0
        for r in ws.iter_rows(min_row=2,values_only=True):
            if isinstance(r[5],(int,float)): total+=r[5]
            if isinstance(r[7],(int,float)): saldo+=r[7]
        return {'total_mes':round(total,2),'saldo_mes':round(saldo,2),'path':path}

try:
    import tkinter as tk
    from tkinter import messagebox,simpledialog,filedialog
    TK=True
except: TK=False

if TK:
    class App:
        def __init__(self,root):
            self.core=PontoCore(); self.root=root; root.title('Batimento de Ponto'); root.geometry('420x380')
            tk.Label(root,text='Usuário:').pack(); self.user=tk.Entry(root); self.user.pack()
            tk.Label(root,text='Senha:').pack(); self.pw=tk.Entry(root,show='*'); self.pw.pack()
            tk.Label(root,text='Horas de trabalho (ex: 8 ou 7.5):').pack(); self.horas=tk.Entry(root); self.horas.pack()
            tk.Button(root,text='Registrar Ponto',command=self.registrar).pack(pady=6)
            tk.Button(root,text='Novo Usuário',command=self.novo_usuario).pack(pady=3)
            tk.Button(root,text='Definir Carga Horária',command=self.definir_horas).pack(pady=3)
            tk.Button(root,text='Fechar Folha',command=self.fechar).pack(pady=6)
            self.status=tk.StringVar(); tk.Label(root,textvariable=self.status,fg='blue').pack(pady=4)
        def registrar(self):
            u,p,h=self.user.get().strip(),self.pw.get().strip(),self.horas.get().strip()
            try:
                path=self.core.register_punch(u,p,h if h else None)
                messagebox.showinfo('OK',f'Registro salvo em:\n{path}')
            except Exception as e: messagebox.showerror('Erro',str(e))
        def novo_usuario(self):
            u=simpledialog.askstring('Usuário','Nome:'); s=simpledialog.askstring('Senha','Senha:',show='*')
            if u and s:
                self.core.register_user(u,s); messagebox.showinfo('OK',f'Usuário {u} criado (7.20h padrão).')
        def definir_horas(self):
            u=self.user.get().strip()
            if not u: messagebox.showerror('Erro','Informe o usuário.'); return
            h=simpledialog.askstring('Horas','Digite a carga diária (ex: 7.5):')
            if h:
                self.core.set_user_hours(u,h); messagebox.showinfo('OK',f'{u}: {h}h por dia.')
        def fechar(self):
            u=self.user.get().strip()
            if not u: messagebox.showerror('Erro','Informe o usuário.'); return
            try:
                r=self.core.close_sheet(u)
                messagebox.showinfo('Resumo',f"Total: {r['total_mes']}h\nSaldo: {r['saldo_mes']}h")
            except Exception as e: messagebox.showerror('Erro',str(e))
    if __name__=='__main__': root=tk.Tk(); App(root); root.mainloop()
else:
    if __name__=='__main__':
        c=PontoCore(); print('Modo console ativo (Tkinter indisponível).')
        while True:
            op=input('\n1) Novo usuário\n2) Registrar ponto\n3) Definir horas do usuário\n4) Fechar folha\n0) Sair\n> ').strip()
            if op=='1':
                u=input('Usuário: '); s=input('Senha: '); c.register_user(u,s); print('Usuário criado (8h padrão).')
            elif op=='2':
                u=input('Usuário: '); s=input('Senha: '); h=input('Horas do dia (vazio=padrão): ')
                try:
                    p=c.register_punch(u,s,h if h else None); print('Ponto registrado em',p)
                except Exception as e: print('Erro:',e)
            elif op=='3':
                u=input('Usuário: '); h=input('Horas por dia: '); c.set_user_hours(u,h); print('Carga atualizada.')
            elif op=='4':
                u=input('Usuário: ')
                try:
                    r=c.close_sheet(u); print('Total',r['total_mes'],'Saldo',r['saldo_mes'])
                except Exception as e: print('Erro:',e)
            elif op=='0': break

